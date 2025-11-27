from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session, flash
from models import db, Order, ProductionStage, TimeLog, User
from datetime import datetime
from functools import wraps
import qrcode
import io
import os
import secrets
from openpyxl import Workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///production.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Use environment variable for SECRET_KEY in production, generate random one for development
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or secrets.token_hex(32)

db.init_app(app)

# ========== Constants for Project Data Validation ==========
VALID_SYSTEMS = ['SLIM', 'JENSEN', 'LITE', 'OTTOSTUM', 'RPTECHNIK', 'W10']
VALID_HANDLE_STYLES = ['1', '2', '3', '4', '5', 'kaseta']


# ========== Authentication Decorators ==========

def login_required(f):
    """Decorator to require login for a route"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Musisz się zalogować aby uzyskać dostęp do tej strony.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def role_required(*roles):
    """Decorator to require specific role(s) for a route"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Musisz się zalogować aby uzyskać dostęp do tej strony.', 'error')
                return redirect(url_for('login'))
            
            user = User.query.get(session['user_id'])
            if not user or user.role not in roles:
                flash('Nie masz uprawnień do tej strony.', 'error')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def get_current_user():
    """Get the currently logged-in user"""
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None

# Create QR codes directory
QR_CODE_DIR = os.path.join(app.root_path, 'static', 'qr_codes')
os.makedirs(QR_CODE_DIR, exist_ok=True)


@app.route('/')
@login_required
def index():
    """Home page with navigation to different panels"""
    user = get_current_user()
    return render_template('index.html', user=user)


# ========== Authentication Routes ==========

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password) and user.is_active:
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash(f'Witaj, {user.full_name}!', 'success')
            
            # Redirect to user's designated panel based on role
            if user.role == 'admin':
                return redirect(url_for('admin_panel'))
            elif user.role == 'designer':
                return redirect(url_for('designer_panel'))
            elif user.role == 'worker':
                return redirect(url_for('worker_panel'))
            elif user.role == 'manager':
                return redirect(url_for('manager_panel'))
            else:
                return redirect(url_for('index'))
        else:
            flash('Nieprawidłowa nazwa użytkownika lub hasło.', 'error')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Logout and clear session"""
    session.clear()
    flash('Zostałeś wylogowany.', 'info')
    return redirect(url_for('login'))


# ========== Admin Panel ==========

@app.route('/admin')
@role_required('admin')
def admin_panel():
    """Admin panel for user and process management"""
    users = User.query.all()
    stages = ProductionStage.query.all()
    return render_template('admin.html', users=users, stages=stages, user=get_current_user())


@app.route('/api/users', methods=['GET', 'POST'])
@role_required('admin')
def manage_users():
    """Get all users or create a new user"""
    if request.method == 'GET':
        users = User.query.all()
        return jsonify([{
            'id': user.id,
            'username': user.username,
            'full_name': user.full_name,
            'role': user.role,
            'is_active': user.is_active,
            'assigned_stages': [{'id': s.id, 'name': s.name} for s in user.assigned_stages]
        } for user in users]), 200
    
    elif request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        full_name = data.get('full_name')
        role = data.get('role')
        stage_ids = data.get('stage_ids', [])
        
        if not all([username, password, full_name, role]):
            return jsonify({'error': 'All fields are required'}), 400
        
        if User.query.filter_by(username=username).first():
            return jsonify({'error': 'Username already exists'}), 400
        
        user = User(username=username, full_name=full_name, role=role)
        user.set_password(password)
        
        # Assign stages if role is worker
        if role == 'worker' and stage_ids:
            stages = ProductionStage.query.filter(ProductionStage.id.in_(stage_ids)).all()
            user.assigned_stages = stages
        
        db.session.add(user)
        db.session.commit()
        
        return jsonify({
            'id': user.id,
            'username': user.username,
            'full_name': user.full_name,
            'role': user.role
        }), 201


@app.route('/api/users/<int:user_id>', methods=['PUT', 'DELETE'])
@role_required('admin')
def manage_user(user_id):
    """Update or delete a user"""
    user = User.query.get_or_404(user_id)
    
    if request.method == 'PUT':
        data = request.json
        
        if 'full_name' in data:
            user.full_name = data['full_name']
        if 'role' in data:
            user.role = data['role']
        if 'is_active' in data:
            user.is_active = data['is_active']
        if 'password' in data and data['password']:
            user.set_password(data['password'])
        if 'stage_ids' in data:
            stage_ids = data['stage_ids']
            stages = ProductionStage.query.filter(ProductionStage.id.in_(stage_ids)).all()
            user.assigned_stages = stages
        
        db.session.commit()
        
        return jsonify({
            'id': user.id,
            'username': user.username,
            'full_name': user.full_name,
            'role': user.role,
            'is_active': user.is_active,
            'assigned_stages': [{'id': s.id, 'name': s.name} for s in user.assigned_stages]
        }), 200
    
    elif request.method == 'DELETE':
        db.session.delete(user)
        db.session.commit()
        return jsonify({'message': 'User deleted'}), 200


# ========== Designer Panel ==========

@app.route('/designer')
@role_required('admin', 'designer')
def designer_panel():
    """Designer panel for creating orders and generating QR codes"""
    orders = Order.query.order_by(Order.created_at.desc()).all()
    return render_template('designer.html', orders=orders, user=get_current_user())


@app.route('/api/orders', methods=['POST'])
def create_order():
    """API endpoint to create a new order"""
    data = request.json
    order_number = data.get('order_number')
    description = data.get('description', '')
    
    # New project data fields
    system = data.get('system')
    handle_style = data.get('handle_style')
    welding_frames_qty = data.get('welding_frames_qty')
    glazing_frames_qty = data.get('glazing_frames_qty')
    szpros_complication = data.get('szpros_complication')
    
    if not order_number:
        return jsonify({'error': 'Order number is required'}), 400
    
    # Validate system value
    if system and system not in VALID_SYSTEMS:
        return jsonify({'error': f'Invalid system. Must be one of: {", ".join(VALID_SYSTEMS)}'}), 400
    
    # Validate handle_style value
    if handle_style and handle_style not in VALID_HANDLE_STYLES:
        return jsonify({'error': f'Invalid handle style. Must be one of: {", ".join(VALID_HANDLE_STYLES)}'}), 400
    
    # Validate welding_frames_qty (1-15)
    if welding_frames_qty is not None:
        if not isinstance(welding_frames_qty, int) or welding_frames_qty < 1 or welding_frames_qty > 15:
            return jsonify({'error': 'Welding frames quantity must be between 1 and 15'}), 400
    
    # Validate glazing_frames_qty (1-15)
    if glazing_frames_qty is not None:
        if not isinstance(glazing_frames_qty, int) or glazing_frames_qty < 1 or glazing_frames_qty > 15:
            return jsonify({'error': 'Glazing frames quantity must be between 1 and 15'}), 400
    
    # Validate szpros_complication (1-5)
    if szpros_complication is not None:
        if not isinstance(szpros_complication, int) or szpros_complication < 1 or szpros_complication > 5:
            return jsonify({'error': 'Szpros complication must be between 1 and 5'}), 400
    
    # Check if order already exists
    existing_order = Order.query.filter_by(order_number=order_number).first()
    if existing_order:
        return jsonify({'error': 'Order number already exists'}), 400
    
    order = Order(
        order_number=order_number,
        description=description,
        system=system,
        handle_style=handle_style,
        welding_frames_qty=welding_frames_qty,
        glazing_frames_qty=glazing_frames_qty,
        szpros_complication=szpros_complication
    )
    db.session.add(order)
    db.session.commit()
    
    return jsonify({
        'id': order.id,
        'order_number': order.order_number,
        'description': order.description,
        'system': order.system,
        'handle_style': order.handle_style,
        'welding_frames_qty': order.welding_frames_qty,
        'glazing_frames_qty': order.glazing_frames_qty,
        'szpros_complication': order.szpros_complication,
        'created_at': order.created_at.isoformat()
    }), 201


@app.route('/api/orders/<int:order_id>/qrcode')
def generate_qr_code(order_id):
    """Generate QR code for an order"""
    order = Order.query.get_or_404(order_id)
    
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(f"ORDER:{order.order_number}")
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to bytes
    img_io = io.BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    
    return send_file(img_io, mimetype='image/png', as_attachment=True, 
                     download_name=f'qr_order_{order.order_number}.png')


# ========== Worker Panel ==========

@app.route('/worker')
@role_required('admin', 'worker')
def worker_panel():
    """Worker panel for scanning QR codes and tracking time"""
    user = get_current_user()
    # Show only assigned stages for workers, all stages for admin
    if user.role == 'worker':
        stages = user.assigned_stages
    else:
        stages = ProductionStage.query.all()
    return render_template('worker.html', stages=stages, user=user)


@app.route('/api/scan', methods=['POST'])
def process_scan():
    """Process QR code scan and start/stop time tracking"""
    data = request.json
    qr_data = data.get('qr_data')
    worker_name = data.get('worker_name')
    stage_id = data.get('stage_id')
    action = data.get('action')  # 'start' or 'stop'
    
    if not all([qr_data, worker_name, stage_id, action]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    # Extract order number from QR code
    if not qr_data.startswith('ORDER:'):
        return jsonify({'error': 'Invalid QR code format'}), 400
    
    order_number = qr_data.replace('ORDER:', '')
    order = Order.query.filter_by(order_number=order_number).first()
    
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    
    stage = ProductionStage.query.get(stage_id)
    if not stage:
        return jsonify({'error': 'Production stage not found'}), 404
    
    if action == 'start':
        # Check if there's already an active session
        active_log = TimeLog.query.filter_by(
            order_id=order.id,
            stage_id=stage_id,
            worker_name=worker_name,
            status='in_progress'
        ).first()
        
        if active_log:
            return jsonify({'error': 'You already have an active session for this order and stage'}), 400
        
        # Create new time log
        time_log = TimeLog(
            order_id=order.id,
            stage_id=stage_id,
            worker_name=worker_name,
            start_time=datetime.utcnow(),
            status='in_progress'
        )
        db.session.add(time_log)
        db.session.commit()
        
        return jsonify({
            'message': 'Work started',
            'log_id': time_log.id,
            'order_number': order.order_number,
            'stage': stage.name,
            'start_time': time_log.start_time.isoformat()
        }), 201
    
    elif action == 'stop':
        # Find active session
        active_log = TimeLog.query.filter_by(
            order_id=order.id,
            stage_id=stage_id,
            worker_name=worker_name,
            status='in_progress'
        ).first()
        
        if not active_log:
            return jsonify({'error': 'No active session found for this order and stage'}), 404
        
        # Stop the session
        active_log.end_time = datetime.utcnow()
        active_log.status = 'completed'
        db.session.commit()
        
        return jsonify({
            'message': 'Work stopped',
            'log_id': active_log.id,
            'order_number': order.order_number,
            'stage': stage.name,
            'duration_minutes': active_log.duration_minutes
        }), 200
    
    else:
        return jsonify({'error': 'Invalid action. Use "start" or "stop"'}), 400


@app.route('/api/worker/active-sessions')
def get_active_sessions():
    """Get all active sessions for a worker"""
    worker_name = request.args.get('worker_name')
    
    if not worker_name:
        return jsonify({'error': 'Worker name is required'}), 400
    
    active_logs = TimeLog.query.filter_by(
        worker_name=worker_name,
        status='in_progress'
    ).all()
    
    sessions = []
    for log in active_logs:
        sessions.append({
            'log_id': log.id,
            'order_number': log.order.order_number,
            'qr_data': f'ORDER:{log.order.order_number}',
            'stage_id': log.stage_id,
            'stage_name': log.stage.name,
            'worker_name': log.worker_name,  # Add worker_name to response
            'start_time': log.start_time.isoformat()
        })
    
    return jsonify(sessions), 200


# ========== Manager Panel ==========

@app.route('/manager')
@role_required('admin', 'manager')
def manager_panel():
    """Manager panel for viewing reports and analytics"""
    orders = Order.query.all()
    stages = ProductionStage.query.all()
    return render_template('manager.html', orders=orders, stages=stages, user=get_current_user())


@app.route('/api/reports/order-times')
def get_order_times_report():
    """Get time report for all orders"""
    order_id = request.args.get('order_id', type=int)
    system = request.args.get('system')
    handle_style = request.args.get('handle_style')
    welding_frames_min = request.args.get('welding_frames_min', type=int)
    glazing_frames_min = request.args.get('glazing_frames_min', type=int)
    szpros_complication = request.args.get('szpros_complication', type=int)
    
    # Calculate duration in days (SQLite Julian day difference)
    duration_days = db.func.julianday(TimeLog.end_time) - db.func.julianday(TimeLog.start_time)
    
    query = db.session.query(
        Order.order_number,
        Order.description,
        Order.system,
        Order.handle_style,
        Order.welding_frames_qty,
        Order.glazing_frames_qty,
        Order.szpros_complication,
        ProductionStage.name.label('stage_name'),
        db.func.count(TimeLog.id).label('work_sessions'),
        db.func.sum(duration_days).label('total_days')
    ).select_from(Order)\
     .join(TimeLog, Order.id == TimeLog.order_id)\
     .join(ProductionStage, TimeLog.stage_id == ProductionStage.id)\
     .filter(TimeLog.status == 'completed')
    
    if order_id:
        query = query.filter(Order.id == order_id)
    if system:
        query = query.filter(Order.system == system)
    if handle_style:
        query = query.filter(Order.handle_style == handle_style)
    if welding_frames_min:
        query = query.filter(Order.welding_frames_qty >= welding_frames_min)
    if glazing_frames_min:
        query = query.filter(Order.glazing_frames_qty >= glazing_frames_min)
    if szpros_complication:
        query = query.filter(Order.szpros_complication == szpros_complication)
    
    query = query.group_by(Order.id, ProductionStage.id)
    
    results = query.all()
    
    report_data = []
    for row in results:
        total_minutes = (row.total_days * 24 * 60) if row.total_days else 0
        report_data.append({
            'order_number': row.order_number,
            'description': row.description,
            'system': row.system,
            'handle_style': row.handle_style,
            'welding_frames_qty': row.welding_frames_qty,
            'glazing_frames_qty': row.glazing_frames_qty,
            'szpros_complication': row.szpros_complication,
            'stage_name': row.stage_name,
            'work_sessions': row.work_sessions,
            'total_minutes': round(total_minutes, 2),
            'total_hours': round(total_minutes / 60, 2)
        })
    
    return jsonify(report_data), 200


@app.route('/api/reports/worker-productivity')
def get_worker_productivity_report():
    """Get productivity report by worker"""
    # Calculate duration in days (SQLite Julian day difference)
    duration_days = db.func.julianday(TimeLog.end_time) - db.func.julianday(TimeLog.start_time)
    
    results = db.session.query(
        TimeLog.worker_name,
        db.func.count(TimeLog.id).label('work_sessions'),
        db.func.sum(duration_days).label('total_days')
    ).filter(TimeLog.status == 'completed')\
     .group_by(TimeLog.worker_name).all()
    
    report_data = []
    for row in results:
        total_minutes = (row.total_days * 24 * 60) if row.total_days else 0
        report_data.append({
            'worker_name': row.worker_name,
            'work_sessions': row.work_sessions,
            'total_minutes': round(total_minutes, 2),
            'total_hours': round(total_minutes / 60, 2)
        })
    
    return jsonify(report_data), 200


@app.route('/api/reports/stage-efficiency')
def get_stage_efficiency_report():
    """Get efficiency report by production stage"""
    # Calculate duration in days (SQLite Julian day difference)
    duration_days = db.func.julianday(TimeLog.end_time) - db.func.julianday(TimeLog.start_time)
    
    results = db.session.query(
        ProductionStage.name,
        db.func.count(TimeLog.id).label('work_sessions'),
        db.func.avg(duration_days).label('avg_days'),
        db.func.sum(duration_days).label('total_days')
    ).select_from(ProductionStage)\
     .join(TimeLog, TimeLog.stage_id == ProductionStage.id)\
     .filter(TimeLog.status == 'completed')\
     .group_by(ProductionStage.id).all()
    
    report_data = []
    for row in results:
        avg_minutes = (row.avg_days * 24 * 60) if row.avg_days else 0
        total_minutes = (row.total_days * 24 * 60) if row.total_days else 0
        report_data.append({
            'stage_name': row.name,
            'work_sessions': row.work_sessions,
            'avg_minutes': round(avg_minutes, 2),
            'avg_hours': round(avg_minutes / 60, 2),
            'total_minutes': round(total_minutes, 2),
            'total_hours': round(total_minutes / 60, 2)
        })
    
    return jsonify(report_data), 200


# ========== Export Reports to XLSX ==========

@app.route('/api/reports/order-times/export')
@role_required('admin', 'manager')
def export_order_times_report():
    """Export order times report to XLSX file"""
    order_id = request.args.get('order_id', type=int)
    system = request.args.get('system')
    handle_style = request.args.get('handle_style')
    welding_frames_min = request.args.get('welding_frames_min', type=int)
    glazing_frames_min = request.args.get('glazing_frames_min', type=int)
    szpros_complication = request.args.get('szpros_complication', type=int)
    
    # Calculate duration in days (SQLite Julian day difference)
    duration_days = db.func.julianday(TimeLog.end_time) - db.func.julianday(TimeLog.start_time)
    
    query = db.session.query(
        Order.order_number,
        Order.description,
        Order.system,
        Order.handle_style,
        Order.welding_frames_qty,
        Order.glazing_frames_qty,
        Order.szpros_complication,
        ProductionStage.name.label('stage_name'),
        db.func.count(TimeLog.id).label('work_sessions'),
        db.func.sum(duration_days).label('total_days')
    ).select_from(Order)\
     .join(TimeLog, Order.id == TimeLog.order_id)\
     .join(ProductionStage, TimeLog.stage_id == ProductionStage.id)\
     .filter(TimeLog.status == 'completed')
    
    if order_id:
        query = query.filter(Order.id == order_id)
    if system:
        query = query.filter(Order.system == system)
    if handle_style:
        query = query.filter(Order.handle_style == handle_style)
    if welding_frames_min:
        query = query.filter(Order.welding_frames_qty >= welding_frames_min)
    if glazing_frames_min:
        query = query.filter(Order.glazing_frames_qty >= glazing_frames_min)
    if szpros_complication:
        query = query.filter(Order.szpros_complication == szpros_complication)
    
    query = query.group_by(Order.id, ProductionStage.id)
    results = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Czasy zleceń"
    
    # Add headers
    headers = ['Zlecenie', 'Opis', 'System', 'Klamka', 'Ramy spaw.', 'Ramy szkl.', 
               'Szprosy', 'Etap', 'Liczba sesji', 'Całkowity czas (min)', 'Całkowity czas (godz)']
    ws.append(headers)
    
    # Add data
    for row in results:
        total_minutes = (row.total_days * 24 * 60) if row.total_days else 0
        ws.append([
            row.order_number,
            row.description or '',
            row.system or '',
            row.handle_style or '',
            row.welding_frames_qty or '',
            row.glazing_frames_qty or '',
            row.szpros_complication or '',
            row.stage_name,
            row.work_sessions,
            round(total_minutes, 2),
            round(total_minutes / 60, 2)
        ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'raport_czasy_zlecen_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/api/reports/worker-productivity/export')
@role_required('admin', 'manager')
def export_worker_productivity_report():
    """Export worker productivity report to XLSX file"""
    # Calculate duration in days (SQLite Julian day difference)
    duration_days = db.func.julianday(TimeLog.end_time) - db.func.julianday(TimeLog.start_time)
    
    results = db.session.query(
        TimeLog.worker_name,
        db.func.count(TimeLog.id).label('work_sessions'),
        db.func.sum(duration_days).label('total_days')
    ).filter(TimeLog.status == 'completed')\
     .group_by(TimeLog.worker_name).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Wydajność pracowników"
    
    # Add headers
    headers = ['Pracownik', 'Liczba sesji', 'Całkowity czas (min)', 'Całkowity czas (godz)']
    ws.append(headers)
    
    # Add data
    for row in results:
        total_minutes = (row.total_days * 24 * 60) if row.total_days else 0
        ws.append([
            row.worker_name,
            row.work_sessions,
            round(total_minutes, 2),
            round(total_minutes / 60, 2)
        ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'raport_wydajnosc_pracownikow_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/api/reports/stage-efficiency/export')
@role_required('admin', 'manager')
def export_stage_efficiency_report():
    """Export stage efficiency report to XLSX file"""
    # Calculate duration in days (SQLite Julian day difference)
    duration_days = db.func.julianday(TimeLog.end_time) - db.func.julianday(TimeLog.start_time)
    
    results = db.session.query(
        ProductionStage.name,
        db.func.count(TimeLog.id).label('work_sessions'),
        db.func.avg(duration_days).label('avg_days'),
        db.func.sum(duration_days).label('total_days')
    ).select_from(ProductionStage)\
     .join(TimeLog, TimeLog.stage_id == ProductionStage.id)\
     .filter(TimeLog.status == 'completed')\
     .group_by(ProductionStage.id).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Efektywność etapów"
    
    # Add headers
    headers = ['Etap', 'Liczba sesji', 'Średni czas (min)', 'Średni czas (godz)', 
               'Całkowity czas (min)', 'Całkowity czas (godz)']
    ws.append(headers)
    
    # Add data
    for row in results:
        avg_minutes = (row.avg_days * 24 * 60) if row.avg_days else 0
        total_minutes = (row.total_days * 24 * 60) if row.total_days else 0
        ws.append([
            row.name,
            row.work_sessions,
            round(avg_minutes, 2),
            round(avg_minutes / 60, 2),
            round(total_minutes, 2),
            round(total_minutes / 60, 2)
        ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'raport_efektywnosc_etapow_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


# ========== Production Stage Management ==========

@app.route('/api/stages', methods=['GET', 'POST'])
def manage_stages():
    """Get all stages or create a new stage"""
    if request.method == 'GET':
        stages = ProductionStage.query.all()
        return jsonify([{
            'id': stage.id,
            'name': stage.name,
            'description': stage.description
        } for stage in stages]), 200
    
    elif request.method == 'POST':
        data = request.json
        name = data.get('name')
        description = data.get('description', '')
        
        if not name:
            return jsonify({'error': 'Stage name is required'}), 400
        
        stage = ProductionStage(name=name, description=description)
        db.session.add(stage)
        db.session.commit()
        
        return jsonify({
            'id': stage.id,
            'name': stage.name,
            'description': stage.description
        }), 201


@app.route('/api/stages/<int:stage_id>', methods=['GET', 'PUT', 'DELETE'])
def manage_stage(stage_id):
    """Get, update, or delete a specific stage"""
    stage = ProductionStage.query.get_or_404(stage_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': stage.id,
            'name': stage.name,
            'description': stage.description,
            'assigned_workers': stage.assigned_users.count()
        }), 200
    
    elif request.method == 'PUT':
        data = request.json
        if 'name' in data:
            name = data['name'].strip() if data['name'] else ''
            if not name:
                return jsonify({'error': 'Nazwa procesu nie może być pusta'}), 400
            if len(name) > 100:
                return jsonify({'error': 'Nazwa procesu może mieć maksymalnie 100 znaków'}), 400
            stage.name = name
        if 'description' in data:
            description = data['description'].strip() if data['description'] else ''
            if len(description) > 500:
                return jsonify({'error': 'Opis może mieć maksymalnie 500 znaków'}), 400
            stage.description = description
        
        db.session.commit()
        
        return jsonify({
            'id': stage.id,
            'name': stage.name,
            'description': stage.description
        }), 200
    
    elif request.method == 'DELETE':
        # Check if stage has time logs
        if stage.time_logs:
            return jsonify({'error': 'Nie można usunąć procesu, który ma powiązane wpisy czasowe'}), 400
        
        db.session.delete(stage)
        db.session.commit()
        return jsonify({'message': 'Proces został usunięty'}), 200


# Initialize database
with app.app_context():
    db.create_all()
    
    # Add new columns to orders table if they don't exist (for existing databases)
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'orders' in inspector.get_table_names():
        existing_columns = [col['name'] for col in inspector.get_columns('orders')]
        new_columns = [
            ('system', 'VARCHAR(20)'),
            ('handle_style', 'VARCHAR(10)'),
            ('welding_frames_qty', 'INTEGER'),
            ('glazing_frames_qty', 'INTEGER'),
            ('szpros_complication', 'INTEGER')
        ]
        for col_name, col_type in new_columns:
            if col_name not in existing_columns:
                with db.engine.connect() as conn:
                    conn.execute(text(f'ALTER TABLE orders ADD COLUMN {col_name} {col_type}'))
                    conn.commit()
                print(f"Added column '{col_name}' to orders table")
    
    # Create default admin user if no users exist
    if User.query.count() == 0:
        admin = User(
            username='admin',
            full_name='Administrator',
            role='admin',
            is_active=True
        )
        admin.set_password('admin123')  # Default password - should be changed after first login
        db.session.add(admin)
        db.session.commit()
        print("Default admin user created: username='admin', password='admin123'")
        print("IMPORTANT: Please change the default password after first login!")
    
    # Create default production stages if they don't exist
    if ProductionStage.query.count() == 0:
        default_stages = [
            ProductionStage(name='Projektowanie', description='Etap projektowania i przygotowania'),
            ProductionStage(name='Cięcie', description='Etap cięcia materiałów'),
            ProductionStage(name='Montaż', description='Etap montażu komponentów'),
            ProductionStage(name='Kontrola jakości', description='Etap kontroli jakości'),
            ProductionStage(name='Pakowanie', description='Etap pakowania produktu')
        ]
        db.session.add_all(default_stages)
        db.session.commit()


if __name__ == '__main__':
    # Only enable debug mode if explicitly set via environment variable
    # For production, set FLASK_ENV=production
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() in ('true', '1', 'yes')
    # Use 0.0.0.0 to allow external connections; for local development only use 127.0.0.1
    host = os.environ.get('FLASK_HOST', '0.0.0.0')
    port = int(os.environ.get('FLASK_PORT', '5000'))
    app.run(debug=debug_mode, host=host, port=port)
